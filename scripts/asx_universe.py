#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import csv as pycsv
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import List, Optional, Dict, Tuple

import requests
from io import BytesIO

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None  # type: ignore

try:
    import pandas as pd  # type: ignore
except Exception:  # pragma: no cover
    pd = None  # type: ignore


ASX_LISTED_COMPANIES_CSV = "https://www.asx.com.au/asx/research/ASXListedCompanies.csv"
ASX_INVESTMENT_PRODUCTS_MONTHLY_REPORT_PAGE = (
    "https://www.asx.com.au/issuers/investment-products/asx-investment-products-monthly-report"
)

MONTHS = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


@dataclass
class CompanyRow:
    code: str
    name: str

    # Back-compat columns (existing universe.csv layout)
    # - For equities from ASXListedCompanies.csv: `sector` holds the ASX "GICS industry group" text.
    # - For investment products from the monthly report: `sector` is "ETF/ETP" and `industry` is product type/sheet name.
    sector: str = ""
    industry: str = ""

    # Enrichment fields (new)
    # Derived from ASX GICS Code Table (ReferencePoint notice 01/23) at the Industry Group tier.
    gics_sector: str = ""
    gics_sector_code: str = ""             # 2-digit string (e.g. "20")
    gics_industry_group: str = ""
    gics_industry_group_code: str = ""     # 8-digit string (e.g. "20100000")
    asset_type: str = ""                   # EQUITY / ETF_ETP / MFUND / UNKNOWN
    source: str = ""                       # row origin


def normalize_code(code: str) -> str:
    code = (code or "").strip().upper()
    # Keep alnum only (ASX codes are alnum; strip spaces, punctuation)
    code = re.sub(r"[^A-Z0-9]", "", code)
    return code


def yahoo_symbol(code: str) -> str:
    c = normalize_code(code)
    return f"{c}.AX" if c and not c.endswith(".AX") else c


# --- GICS enrichment (Industry Group -> Sector + codes) -----------------------
# ASXListedCompanies.csv only provides the "GICS industry group" (2nd tier).
# We derive the GICS Sector (1st tier) + the official 8-digit Industry Group code
# using ASX Information Services Notice 01/23 attachment:
# "GICS Code Table (source: ASX ReferencePoint Master List Specification Manual)".
#
# Notes:
# - Some labels in ASXListedCompanies.csv are abbreviated (e.g. "Not Applic", "Class Pend")
#   or use "&" instead of "and". We normalize to match.
# - If ASX introduces a new/renamed industry group, the universe build will warn and
#   leave the derived fields blank (so your pipeline doesn't silently lie).

def _norm_gics_label(s: str) -> str:
    s = (s or "").strip().lower()
    # normalize ampersands and punctuation
    s = s.replace("&", "and")
    s = re.sub(r"[()]", " ", s)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

# Mapping keyed by normalized Industry Group label -> (Sector label, Industry Group code)
# (The sector code is the first two digits of the Industry Group code.)
_GICS_IG_TO_SECTOR_AND_CODE: Dict[str, Tuple[str, str]] = {
    # Energy (10)
    _norm_gics_label("Energy"): ("Energy", "10100000"),

    # Materials (15)
    _norm_gics_label("Materials"): ("Materials", "15100000"),

    # Industrials (20)
    _norm_gics_label("Capital Goods"): ("Industrials", "20100000"),
    _norm_gics_label("Commercial & Professional Services"): ("Industrials", "20200000"),
    _norm_gics_label("Transportation"): ("Industrials", "20300000"),

    # Consumer Discretionary (25)
    _norm_gics_label("Automobile & Components"): ("Consumer Discretionary", "25100000"),
    _norm_gics_label("Automobiles & Components"): ("Consumer Discretionary", "25100000"),
    _norm_gics_label("Consumer Durables & Apparel"): ("Consumer Discretionary", "25200000"),
    _norm_gics_label("Consumer Services"): ("Consumer Discretionary", "25300000"),
    _norm_gics_label("Consumer Discretionary Distribution & Retail"): ("Consumer Discretionary", "25500000"),

    # Consumer Staples (30)
    _norm_gics_label("Consumer Staples Distribution & Retail"): ("Consumer Staples", "30100000"),
    _norm_gics_label("Food, Beverage & Tobacco"): ("Consumer Staples", "30200000"),
    _norm_gics_label("Food Beverage & Tobacco"): ("Consumer Staples", "30200000"),
    _norm_gics_label("Household & Personal Products"): ("Consumer Staples", "30300000"),

    # Health Care (35)
    _norm_gics_label("Health Care Equipment & Services"): ("Health Care", "35100000"),
    _norm_gics_label("Pharmaceuticals, Biotechnology & Life Sciences"): ("Health Care", "35200000"),

    # Financials (40)
    _norm_gics_label("Banks"): ("Financials", "40100000"),
    _norm_gics_label("Financial Services"): ("Financials", "40200000"),
    _norm_gics_label("Insurance"): ("Financials", "40300000"),

    # Information Technology (45)
    _norm_gics_label("Software & Services"): ("Information Technology", "45100000"),
    _norm_gics_label("Technology Hardware & Equipment"): ("Information Technology", "45200000"),
    _norm_gics_label("Semiconductors & Semiconductor Equipment"): ("Information Technology", "45300000"),

    # Communication Services (50)
    _norm_gics_label("Telecommunication Services"): ("Communication Services", "50100000"),
    _norm_gics_label("Media and Entertainment"): ("Communication Services", "50200000"),
    _norm_gics_label("Media & Entertainment"): ("Communication Services", "50200000"),

    # Utilities (55)
    _norm_gics_label("Utilities"): ("Utilities", "55100000"),

    # Real Estate (60)
    _norm_gics_label("Equity Real Estate Investment Trusts REITs"): ("Real Estate", "60100000"),
    _norm_gics_label("Equity Real Estate Investment Trusts (REITs)"): ("Real Estate", "60100000"),
    _norm_gics_label("Real Estate Management & Development"): ("Real Estate", "60200000"),

    # Special / internal
    _norm_gics_label("ASX Internal Classification Pending"): ("Classification Pending", "99999999"),
    _norm_gics_label("Class Pend"): ("Classification Pending", "99999999"),
    _norm_gics_label("Classification Pending"): ("Classification Pending", "99999999"),
    _norm_gics_label("Not Applicable"): ("Not Applicable", "00000000"),
    _norm_gics_label("Not Applic"): ("Not Applicable", "00000000"),
}

def derive_gics_from_industry_group(industry_group: str) -> Tuple[str, str, str]:
    """Return (gics_sector, gics_sector_code, gics_industry_group_code) for a given industry group label."""
    ig = (industry_group or "").strip()
    if not ig:
        return ("", "", "")
    key = _norm_gics_label(ig)
    rec = _GICS_IG_TO_SECTOR_AND_CODE.get(key)
    if not rec:
        # unknown group — don't guess
        return ("", "", "")
    sector, ig_code = rec
    sector_code = ig_code[:2] if ig_code and len(ig_code) >= 2 else ""
    return (sector, sector_code, ig_code)


def _http_get(url: str, timeout: int = 30) -> requests.Response:
    """HTTP GET with headers that the ASX site is happy with.

    Note: ASX occasionally blocks non-browser user agents for XLSX/PDF assets.
    We use a mainstream UA string + basic accept headers to reduce 403/400.
    """
    headers = {
        # Browser-like UA; ASX has been known to block non-browser UA strings.
        "User-Agent": (
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "*/*",
        "Accept-Language": "en-AU,en;q=0.9",
        "Connection": "keep-alive",
    }

    # For DAM-hosted assets, a Referer can help (the ASX site can be picky).
    if (
        "asx-investment-products-monthly-report" in url
        or "asx-investment-products-reports" in url
        or "asx-investment-products" in url
    ):
        headers["Referer"] = ASX_INVESTMENT_PRODUCTS_MONTHLY_REPORT_PAGE

    resp = requests.get(url, timeout=timeout, headers=headers)
    resp.raise_for_status()
    return resp


def _find_header_start(lines: List[str]) -> int:
    # ASX CSV sometimes has metadata lines before the real header.
    for i, line in enumerate(lines):
        if line.lower().startswith("company name,"):
            return i
    return 0


def read_asx_listed_companies() -> List[CompanyRow]:
    """Equity universe from ASX 'ASXListedCompanies.csv' (companies only)."""
    resp = _http_get(ASX_LISTED_COMPANIES_CSV)
    raw_lines = resp.text.splitlines()
    header_idx = _find_header_start(raw_lines)
    normalized = "\n".join(raw_lines[header_idx:])

    rows: List[CompanyRow] = []

    if pd is not None:
        from io import StringIO

        df = pd.read_csv(StringIO(normalized))
        rows_iter = df.to_dict(orient="records")
        col_map = {str(c).strip().lower(): c for c in df.columns}
    else:
        rows_iter = list(pycsv.DictReader(normalized.splitlines()))
        col_map = {str(c).strip().lower(): c for c in (rows_iter[0].keys() if rows_iter else [])}

    def find_col(cands: List[str]) -> Optional[str]:
        for cand in cands:
            if cand in col_map:
                return col_map[cand]
        return None

    col_name = find_col(["company name", "name"])
    col_code = find_col(["asx code", "asxcode", "code"])
    col_sector = find_col(["gics industry group", "gics sector", "sector"])
    col_industry = find_col(["gics industry", "industry"])

    for r in rows_iter:
        code = normalize_code(str(r.get(col_code, "") if col_code else ""))
        name = str(r.get(col_name, "") if col_name else "").strip()
        industry_group = str(r.get(col_sector, "") if col_sector else "").strip()

        industry = str(r.get(col_industry, "") if col_industry else "").strip()


        gics_sector, gics_sector_code, gics_ig_code = derive_gics_from_industry_group(industry_group)

        if industry_group and not gics_sector:

            print(f"[warn] unknown GICS industry group '{industry_group}' for {code}", file=sys.stderr)


        if not code:

            continue

        rows.append(

            CompanyRow(

                code=code,

                name=name,

                sector=industry_group,

                industry=industry,

                gics_sector=gics_sector,

                gics_sector_code=gics_sector_code,

                gics_industry_group=industry_group,

                gics_industry_group_code=gics_ig_code,

                asset_type="EQUITY",

                source="ASXListedCompanies.csv",

            )

        )

    return rows


def _score_xlsx_url(url: str) -> Tuple[int, int, int]:
    """Heuristic score (year, month, day) to pick the latest XLSX link."""
    u = url.lower()
    year = 0
    month = 0
    day = 0

    m = re.search(r"/(20\d{2})/", u)
    if m:
        year = int(m.group(1))

    m2 = re.search(r"-(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*-(20\d{2})", u)
    if m2:
        month = MONTHS.get(m2.group(1), 0)
        year = int(m2.group(2))
    else:
        m3 = re.search(r"(20\d{2})[-_/](\d{1,2})", u)
        if m3:
            year = int(m3.group(1))
            month = int(m3.group(2))

    mday = re.search(r"-(\d{1,2})-(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*-(20\d{2})", u)
    if mday:
        day = int(mday.group(1))
        month = MONTHS.get(mday.group(2), month)
        year = int(mday.group(3))

    return (year, month, day)


def _extract_investment_products_xlsx_urls(page_html: str) -> List[str]:
    """Return a de-duplicated list of absolute XLSX URLs found on the page."""
    hrefs = re.findall(r'href="([^"]+\.xlsx)"', page_html, flags=re.IGNORECASE)
    if not hrefs:
        hrefs = re.findall(r"href='([^']+\.xlsx)'", page_html, flags=re.IGNORECASE)

    urls: List[str] = []
    for h in hrefs:
        if h.startswith("//"):
            url = "https:" + h
        elif h.startswith("http"):
            url = h
        else:
            url = "https://www.asx.com.au" + (h if h.startswith("/") else "/" + h)
        urls.append(url)

    # De-dupe preserving order
    seen = set()
    out: List[str] = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            out.append(u)
    return out


def _pick_latest_investment_products_xlsx_urls(page_html: str) -> List[str]:
    """Pick the latest month of report XLSX URLs.

    The ASX page sometimes publishes multiple XLSX variants (e.g., ABS vs non-ABS).
    We select the latest (year, month, day) score and return *all* URLs for that
    same score so we can union the tickers.
    """
    urls = _extract_investment_products_xlsx_urls(page_html)
    if not urls:
        return []

    scored = [(u, _score_xlsx_url(u)) for u in urls]
    scored.sort(key=lambda x: x[1], reverse=True)
    best_score = scored[0][1]

    best = [u for (u, s) in scored if s == best_score]

    # Prefer ABS first (if present) but keep all.
    best.sort(key=lambda u: (0 if u.lower().endswith("abs.xlsx") else 1, u))
    return best


def _http_get_with_retries(url: str, timeout: int = 30, retries: int = 3, backoff_s: float = 1.2) -> requests.Response:
    last_err: Optional[Exception] = None
    for i in range(retries):
        try:
            return _http_get(url, timeout=timeout)
        except Exception as e:  # pragma: no cover
            last_err = e
            if i < retries - 1:
                import time

                time.sleep(backoff_s * (i + 1))
    assert last_err is not None
    raise last_err


def _fallback_non_abs(url: str) -> Optional[str]:
    # common pattern: ...-abs.xlsx
    if re.search(r"-abs\.xlsx$", url, flags=re.IGNORECASE):
        return re.sub(r"-abs\.xlsx$", ".xlsx", url, flags=re.IGNORECASE)
    return None


def _norm_header(v: object) -> str:
    s = "" if v is None else str(v)
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _find_header_row_and_cols(ws) -> Optional[Tuple[int, Dict[str, int]]]:
    """Find a header row and return (row_index_1based, mapping header->col_index_1based)."""
    # Scan first N rows for a row that looks like it contains 'ASX code'.
    # Some monthly reports have a few note rows before the real header.
    max_scan = min(80, ws.max_row or 0)
    for r in range(1, max_scan + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, (ws.max_column or 0) + 1)]
        normed = [_norm_header(v) for v in values]
        joined = " | ".join([x for x in normed if x])
        if "asx code" in joined or "asx code/apir" in joined or "asx code / apir" in joined:
            mapping: Dict[str, int] = {}
            for idx, h in enumerate(normed, start=1):
                if h:
                    mapping[h] = idx
            return (r, mapping)
    return None


def _pick_col(mapping: Dict[str, int], candidates: List[str], contains: List[str] | None = None) -> Optional[int]:
    keys = list(mapping.keys())
    # Exact-ish matches first
    for cand in candidates:
        c = cand.strip().lower()
        if c in mapping:
            return mapping[c]
    # Substring matches
    for k in keys:
        for cand in candidates:
            if cand.strip().lower() in k:
                return mapping[k]
    # Contains-all token match
    if contains:
        want = [t.lower() for t in contains]
        for k in keys:
            if all(t in k for t in want):
                return mapping[k]
    return None


def _parse_investment_products_xlsx_bytes(xlsx_bytes: bytes) -> List[CompanyRow]:
    if openpyxl is None:
        raise RuntimeError("openpyxl not available")

    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)
    rows: List[CompanyRow] = []

    code_candidates = [
        "asx code",
        "asx code / apir code",
        "asx code/apir code",
        "asx code/apir",
        "asxcode",
        "code",
    ]
    name_candidates = [
        "investment name",
        "product name",
        "fund name",
        "security name",
        "name",
        "description",
    ]
    type_candidates = [
        "product type",
        "product",
        "structure",
        "category",
        "asset class",
    ]

    for ws in wb.worksheets:
        found = _find_header_row_and_cols(ws)
        if not found:
            continue
        header_row, mapping = found

        code_col = _pick_col(mapping, code_candidates, contains=["asx", "code"])
        if not code_col:
            continue
        name_col = _pick_col(mapping, name_candidates)
        type_col = _pick_col(mapping, type_candidates)

        for r in range(header_row + 1, (ws.max_row or 0) + 1):
            raw_code = ws.cell(row=r, column=code_col).value
            code = normalize_code("" if raw_code is None else str(raw_code))
            if not code:
                continue

            # Filter out APIR-only rows that have no ASX code (APIR codes are long).
            if len(code) > 7:
                continue

            raw_name = ws.cell(row=r, column=name_col).value if name_col else None
            name = ("" if raw_name is None else str(raw_name)).strip() or "ASX investment product"

            raw_type = ws.cell(row=r, column=type_col).value if type_col else None
            ptype = ("" if raw_type is None else str(raw_type)).strip()

            sector = "ETF/ETP"
            industry = (ptype or ws.title).strip()

            ptype_l = (ptype or "").lower().replace("-", "").replace(" ", "")
            asset_type = "MFUND" if "mfund" in ptype_l else "ETF_ETP"

            rows.append(
                CompanyRow(
                    code=code,
                    name=name,
                    sector=sector,
                    industry=industry,
                    asset_type=asset_type,
                    source="ASX Investment Products Monthly Report",
                )
            )

    # de-dupe by code
    out: Dict[str, CompanyRow] = {}
    for r in rows:
        if r.code and r.code not in out:
            out[r.code] = r
    return list(out.values())


def read_asx_investment_products() -> List[CompanyRow]:
    """ETF/ETP universe from the ASX 'Investment Products Monthly Report'."""
    try:
        page = _http_get_with_retries(ASX_INVESTMENT_PRODUCTS_MONTHLY_REPORT_PAGE, timeout=45).text
        urls = _pick_latest_investment_products_xlsx_urls(page)
        if not urls:
            print("[warn] Could not find XLSX link on monthly report page; skipping ETP expansion", file=sys.stderr)
            return []

        all_rows: Dict[str, CompanyRow] = {}

        for xlsx_url in urls:
            try:
                # Prefer XLSX-appropriate accept header on this request
                resp = _http_get_with_retries(xlsx_url, timeout=90)
                xlsx_bytes = resp.content
            except Exception:
                # Try the non-ABS variant if the ABS asset is blocked/unavailable.
                alt = _fallback_non_abs(xlsx_url)
                if not alt:
                    continue
                try:
                    xlsx_bytes = _http_get_with_retries(alt, timeout=90).content
                except Exception:
                    continue

            try:
                parsed = _parse_investment_products_xlsx_bytes(xlsx_bytes)
            except Exception as e:
                print(f"[warn] Failed to parse investment products XLSX ({xlsx_url}): {e}", file=sys.stderr)
                continue

            for r in parsed:
                if r.code and r.code not in all_rows:
                    all_rows[r.code] = r

        if not all_rows:
            print("[warn] investment products expansion produced 0 tickers (download/parse likely failed)", file=sys.stderr)
        return list(all_rows.values())

    except Exception as e:
        print(f"[warn] investment products expansion failed: {e}", file=sys.stderr)
        return []


def _read_extra_tickers(extra_path: Path) -> List[str]:
    if not extra_path.exists():
        return []
    out: List[str] = []
    for line in extra_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        s = line.strip()
        if not s or s.startswith("#"):
            continue
        if s.upper().endswith(".AX"):
            out.append(s.upper())
        else:
            out.append(yahoo_symbol(s))
    return out


def write_universe_csv(rows: List[CompanyRow], out_csv: Path) -> None:
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    now_dt = datetime.now(timezone.utc).replace(microsecond=0)
    now_utc = now_dt.isoformat().replace("+00:00", "Z")
    awst = now_dt.astimezone(ZoneInfo("Australia/Perth")).strftime("%Y-%m-%d %H:%M %Z")
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = pycsv.writer(f)
        w.writerow(["code","yahoo_symbol","name","sector","industry","last_extracted_utc","last_extracted_awst","gics_sector","gics_sector_code","gics_industry_group","gics_industry_group_code","asset_type","source"])
        for r in rows:
            w.writerow([r.code, yahoo_symbol(r.code), r.name, r.sector, r.industry, now_utc, awst, r.gics_sector, r.gics_sector_code, r.gics_industry_group, r.gics_industry_group_code, r.asset_type, r.source])



def write_universe_json(rows: List[CompanyRow], out_json: Path) -> None:
    """App-friendly universe payload (stable keys, easy to join on yahoo_symbol)."""
    out_json.parent.mkdir(parents=True, exist_ok=True)
    now_dt = datetime.now(timezone.utc).replace(microsecond=0)
    now_utc = now_dt.isoformat().replace("+00:00", "Z")
    awst = now_dt.astimezone(ZoneInfo("Australia/Perth")).strftime("%Y-%m-%d %H:%M %Z")

    payload = {
        "dataset": "asx/universe",
        "asOfUtc": now_utc,
        "asOfPerth": awst,
        "count": len(rows),
        "rows": [
            {
                "code": r.code,
                "yahoo_symbol": yahoo_symbol(r.code),
                "name": r.name,
                # Back-compat / legacy fields
                "sector": r.sector,
                "industry": r.industry,
                # Enrichment
                "gics_sector": r.gics_sector,
                "gics_sector_code": r.gics_sector_code,
                "gics_industry_group": r.gics_industry_group,
                "gics_industry_group_code": r.gics_industry_group_code,
                "asset_type": r.asset_type,
                "source": r.source,
            }
            for r in rows
        ],
    }
    out_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

def write_tickers_txt(rows: List[CompanyRow], out_txt: Path) -> None:
    out_txt.parent.mkdir(parents=True, exist_ok=True)
    now_dt = datetime.now(timezone.utc).replace(microsecond=0)
    now_utc = now_dt.isoformat().replace("+00:00", "Z")
    awst = now_dt.astimezone(ZoneInfo("Australia/Perth")).strftime("%Y-%m-%d %H:%M %Z")

    tickers = {yahoo_symbol(r.code) for r in rows if r.code}
    tickers |= set(_read_extra_tickers(out_txt.parent / "tickers_extra.txt"))

    tickers_sorted = sorted({t for t in tickers if t})

    lines = [
        "# ASX universe tickers (Yahoo symbols, .AX)",
        f"# last_extracted_utc: {now_utc}",
        f"# last_extracted_awst: {awst}",
        "# one symbol per line",
        *tickers_sorted,
    ]
    out_txt.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out-csv", default="asx/universe.csv")
    ap.add_argument("--out-tickers", default="asx/tickers_asx.txt")
    ap.add_argument("--out-json", default="asx/universe_latest.json")
    args = ap.parse_args()

    rows = read_asx_listed_companies()
    etps = read_asx_investment_products()

    merged: Dict[str, CompanyRow] = {r.code: r for r in rows if r.code}
    for r in etps:
        if r.code and r.code not in merged:
            merged[r.code] = r

    final = list(merged.values())
    final.sort(key=lambda r: r.code)

    write_universe_csv(final, Path(args.out_csv))
    if args.out_json:
        write_universe_json(final, Path(args.out_json))
    write_tickers_txt(final, Path(args.out_tickers))
    print(f"[ok] wrote {args.out_csv}, {args.out_tickers}, {args.out_json} ({len(final)} tickers)")


if __name__ == "__main__":
    main()
